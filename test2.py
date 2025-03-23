# -*- coding: utf-8 -*-
import streamlit as st
from streamlit_extras.add_vertical_space import add_vertical_space
from streamlit_quill import st_quill

# 기존 임포트 아래에 추가
from langchain_community.vectorstores import FAISS
from langchain_openai import OpenAIEmbeddings
from langchain_core.runnables import RunnablePassthrough
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
from langchain_core.output_parsers import StrOutputParser, JsonOutputParser
from dotenv import load_dotenv
from functools import lru_cache
import pyperclip
import os

import pandas as pd
from fpdf import FPDF
import tempfile
from openpyxl import load_workbook
from openpyxl.styles.borders import Border

import json
from bs4 import BeautifulSoup


# API KEY 정보로드
load_dotenv()

llm = ChatOpenAI(model_name="gpt-4o-mini", temperature=0)

embeddings = OpenAIEmbeddings(
    model="text-embedding-3-large",
)
embedding_path = "AI-Streamlit/pages/curriculum_keyword_vectorstore"

vectorstore = FAISS.load_local(
    embedding_path, 
    embeddings,
    allow_dangerous_deserialization=True  # 신뢰할 수 있는 경우에만 설정
)

import streamlit.components.v1 as components
import json

import streamlit.components.v1 as components
import json

def copy_to_clipboard_js(text):
    js_code = f"""
    <html>
      <body>
        <script>
          // fallback: document.execCommand("copy")
          function fallbackCopyTextToClipboard(text) {{
            var textArea = document.createElement("textarea");
            textArea.value = text;
            // 최소한 화면에 표시되지 않도록 스타일 지정
            textArea.style.position = "fixed";
            textArea.style.top = "0";
            textArea.style.left = "0";
            textArea.style.width = "2em";
            textArea.style.height = "2em";
            textArea.style.padding = "0";
            textArea.style.border = "none";
            textArea.style.outline = "none";
            textArea.style.boxShadow = "none";
            textArea.style.background = "transparent";
            document.body.appendChild(textArea);
            textArea.focus();
            textArea.select();
            try {{
              var successful = document.execCommand('copy');
              console.log("Fallback: Copying text command was " + (successful ? "successful" : "unsuccessful"));
            }} catch (err) {{
              console.error("Fallback: Unable to copy", err);
            }}
            document.body.removeChild(textArea);
          }}

          function copyText(text) {{
            // 시도: document와 window에 강제로 포커스 요청
            try {{
              window.focus();
              document.body.focus();
            }} catch (e) {{
              console.warn("Could not focus document:", e);
            }}

            if (navigator.clipboard && navigator.clipboard.writeText) {{
              navigator.clipboard.writeText(text).then(function() {{
                  console.log("Copied to clipboard successfully!");
              }}, function(err) {{
                  console.error("Could not copy text: ", err);
                  fallbackCopyTextToClipboard(text);
              }});
            }} else {{
              fallbackCopyTextToClipboard(text);
            }}
          }}

          copyText({json.dumps(text)});
        </script>
      </body>
    </html>
    """
    # height와 width를 0으로 설정하여 화면에 보이지 않게 합니다.
    components.html(js_code, height=0, width=0)


def format_copied_text(html_content):
    """
    st_quill 위젯에서 얻은 HTML 문자열을 깔끔한 텍스트로 변환합니다.
    
    예시)
    <p><br></p><p>                    1. 주제</p><p>                    국회에서 하는 일</p>
    =>
    1. 주제

    국회에서 하는 일
    """
    from bs4 import BeautifulSoup  # 이미 상단에 import 되어 있다면 생략 가능
    # HTML 파싱 및 텍스트 추출
    soup = BeautifulSoup(html_content, "html.parser")
    # <br> 태그와 <p> 태그 사이에 개행 문자를 추가하여 텍스트 추출
    text = soup.get_text(separator="\n", strip=True)
    # 줄 단위로 분리한 후, 빈 줄은 제거
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    # 단락(예: <p> 태그 단위)을 두 줄 띄워서 구분
    formatted_text = "\n\n".join(lines)
    return formatted_text


def parse_report_content(raw_content):
    """
    st_quill 에디터에서 받아온 raw_content(문자열 또는 HTML)를 파싱하여
    {"주제": ..., "조사 개요": ..., "조사 내용": ..., "출처": ...} 형태의 딕셔너리로 변환합니다.
    """
    # 이미 딕셔너리면 그대로 반환
    if isinstance(raw_content, dict):
        return raw_content

    # 우선 JSON 파싱 시도
    try:
        return json.loads(raw_content)
    except Exception:
        pass

    # JSON 파싱에 실패하면, HTML 태그 제거 후 간단한 텍스트 분리 시도
    soup = BeautifulSoup(raw_content, "html.parser")
    text = soup.get_text(separator="\n")
    sections = {}
    current_key = None
    for line in text.splitlines():
        line = line.strip()
        if line.startswith("1. 주제"):
            current_key = "주제"
            sections[current_key] = ""
        elif line.startswith("2. 조사 개요"):
            current_key = "조사 개요"
            sections[current_key] = ""
        elif line.startswith("3. 조사 내용"):
            current_key = "조사 내용"
            sections[current_key] = ""
        elif line.startswith("4. 출처"):
            current_key = "출처"
            sections[current_key] = ""
        else:
            if current_key:
                sections[current_key] += line + "\n"
    return sections

#@st.cache_data(max_entries=32)
def query_keywords(vectorstore, subject, grade, topic):
    """학년, 교과, 주제어 입력 시 해당 키워드 검색"""
    query_text = f"교과: {subject} 학년: {grade} 주제어: {topic}"
    results = vectorstore.similarity_search(query_text, k=1)
    
    if not results:
        return []
    
    best_match = results[0]
    return best_match.metadata['keywords']

from bs4 import BeautifulSoup

def parse_report_html_to_json(html_content):
    # 1. HTML 태그를 제거하고 텍스트로 변환
    soup = BeautifulSoup(html_content, "html.parser")
    # 각 태그 사이에 개행문자를 넣어 텍스트를 깔끔하게 가져옵니다.
    text = soup.get_text(separator="\n", strip=True)
    
    # 2. 마커 정의
    markers = ["1. 주제", "2. 조사 개요", "3. 조사 내용", "4. 출처"]
    result = {}
    
    for i, marker in enumerate(markers):
        start = text.find(marker)
        if start == -1:
            # 마커가 없다면 빈 문자열로 처리
            result[marker] = ""
            continue
        # 마커 바로 뒤부터 내용을 추출 (마커 문자열 길이만큼 이동)
        start += len(marker)
        # 다음 마커가 있다면 그 전까지의 내용을 추출, 없으면 텍스트 끝까지
        if i < len(markers) - 1:
            end = text.find(markers[i+1])
            if end == -1:
                end = len(text)
        else:
            end = len(text)
        section_text = text[start:end].strip()
        
        # 3. 키 변환: 
        # "1. 주제" → "주제(키워드)" 나머지는 마커의 숫자와 점을 제거한 값
        if marker.startswith("1."):
            key = "주제(키워드)"
        else:
            key = marker.split(".", 1)[1].strip()
        result[key] = section_text
    
    # 4. 추가 키: "알게 된 점 또는 느낀 점"은 빈 문자열로 설정
    result["알게 된 점 또는 느낀 점"] = ""
    
    return result



@st.cache_data(max_entries=32)
def keywords_recomand_rag(vectorstore, subject, grade, topic):
    return query_keywords(vectorstore, subject, grade, topic)

# 주제 생성 함수
@st.cache_data(max_entries=32)
def generate_topics(grade, selected_subject, keyword):
    return query_keywords(vectorstore, selected_subject, grade, keyword)

#@st.cache_data(max_entries=32)
def create_report_prompt(grade, selected_subject, topic):
    prompt_template = f"""
    초등학생 {grade}학년 수준에서 "{selected_subject}" 과목의 "{topic}"에 대한 학습용 연구 보고서를 총 1,500자 수준으로 작성해주세요.

    [보고서 형식]
    1. 주제 (키워드)
    2. 조사 개요
    3. 조사 내용
    4. 출처

    위 형식을 참고하여 보고서를 작성하되, 각각의 값들을 JSON으로 받아와주세요.
    주어진 형식과 동일한 JSON 형태로만 응답하세요. JSON 외의 추가 텍스트는 출력하지 마세요.
    """
    return ChatPromptTemplate.from_template(prompt_template)

#@st.cache_data(max_entries=32)
def generate_question(grade, selected_subject, topic):
    prompt = create_question_prompt(grade, selected_subject, topic)
    # Build a simple chain: prompt | llm | output parser
    chain = prompt | llm | JsonOutputParser()
    response = chain.invoke({})
    return response

    #@st.cache_data(max_entries=32)

#@st.cache_data(max_entries=32)
def generate_report(grade, selected_subject, topic):
    prompt = create_report_prompt(grade, selected_subject, topic)
    # Build a simple chain: prompt | llm | output parser
    chain = prompt | llm | JsonOutputParser()
    response = chain.invoke({})
    return response

#@st.cache_data(max_entries=32)
# def create_summary_prompt():
#     prompt_template = """
#     다음은 연구 보고서의 내용입니다. 이 내용을 간결하게 요약해 주세요. 요약은 다음 형식을 따르세요:
#     [요약 형식]
#     1. 각각의 소제목은 JSON 형태의 key로 유지해야합니다.

#     주제: {주제}
#     조사 개요: {조사_개요}
#     조사 내용: {조사_내용}
#     출처: {출처}

#     위 형식을 참고하여 보고서를 작성하되, JSON key는 위에 4개만 있으면, 됩니다 그 하위에 있는 내용은 string으로 작성하세요.
#     JSON 외의 추가 텍스트는 출력하지 마세요.
#     """
#     return ChatPromptTemplate.from_template(prompt_template)

def convert_to_string(content):
    """
    content가 딕셔너리이면 지정된 순서대로 key와 value를 문자열로 결합하고,
    문자열이면 그대로 반환합니다.
    """
    if isinstance(content, dict):
        # 원하는 key 순서 (존재하지 않는 key는 무시됩니다)
        keys_order = ["주제", "조사 개요", "조사 내용", "출처"]
        parts = []
        for key in keys_order:
            if key in content:
                parts.append(f"{key}: {content[key]}")
        # 추가로 포함할 다른 key가 있다면 (선택사항)
        for key, value in content.items():
            if key not in keys_order:
                parts.append(f"{key}: {value}")
        return "\n".join(parts)
    elif isinstance(content, str):
        return content
    else:
        return str(content)


def create_summary_prompt():
    prompt_template = """
    다음은 연구 보고서의 내용입니다. 이 내용을 간결하게 요약해 주세요. 요약은 다음 형식을 따르세요:
    [요약 형식]
    1. 각각의 소제목은 JSON 형태의 key로 유지해야 합니다.

    본문:
    {본문}

    위 형식을 참고하여 보고서를 작성하되, JSON key는 주제, 조사 개요, 조사 내용, 출처 4개만 있으면 됩니다.
    그 하위의 내용은 문자열(string)로 작성하세요.
    JSON 외의 추가 텍스트는 출력하지 마세요.
    """
    return ChatPromptTemplate.from_template(prompt_template)

#@st.cache_data(max_entries=32)
def generate_summary(content):
    # content가 딕셔너리든 문자열이든 문자열로 변환
    text = convert_to_string(content)
    prompt = create_summary_prompt()
    chain = prompt | llm | JsonOutputParser()
    # 프롬프트의 {본문} 플레이스홀더에 변환된 텍스트를 넣어줍니다.
    response = chain.invoke({"본문": text})
    return response

def load_css():
    """CSS 파일들 로드"""
    css_files = ["base", "sidebar", "main"]
    for css_file in css_files:
        with open(f"assets/style_{css_file}.css", encoding='utf-8') as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
            
            
def is_merged_cell(ws, coord):
    """
    주어진 좌표(coord)가 병합된 셀 범위 내에 있는지 확인하고,
    해당 병합 범위 객체를 반환합니다.
    병합된 범위가 아니면 None을 반환합니다.
    """
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return merged_range
    return None

def split_text_lines(pdf, text, w):
    """
    주어진 텍스트(text)를 셀 너비(w) 내에 맞게 여러 줄로 분할하는 함수.
    단어 단위로 분할하여 각 줄의 길이가 w를 넘지 않도록 합니다.
    """
    # 줄바꿈 문자를 보존하기 위해 "\n"을 구분자로 분리
    words = text.replace("\n", " \n ").split()
    lines = []
    current_line = ""
    for word in words:
        if word == "\n":
            lines.append(current_line)
            current_line = ""
        else:
            if current_line == "":
                test_line = word
            else:
                test_line = current_line + " " + word
            if pdf.get_string_width(test_line) <= w:
                current_line = test_line
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
    if current_line:
        lines.append(current_line)
    return lines

def draw_investigation_content_row(pdf, left_text, right_text, left_width, right_width, line_height):
    """
    "조사 내용" 행은 왼쪽 셀에는 헤더(예, "조사 내용")를,
    오른쪽 셀에는 긴 content(여기서는 값 "3")를 출력합니다.
    
    오른쪽 셀의 텍스트가 여러 페이지에 걸쳐 출력되더라도,
    각 페이지에 해당 영역에 대해 테두리를 반드시 그립니다.
    
    - left_width: 왼쪽 셀 너비 (mm)
    - right_width: 오른쪽 셀 너비 (mm)
    - line_height: 기본 줄 높이 (mm)
    """
    x_left = pdf.get_x()
    y_start = pdf.get_y()
    x_right = x_left + left_width

    # 오른쪽 셀 텍스트를 지정된 너비에 맞게 분할
    lines = split_text_lines(pdf, right_text, right_width)
    printed_left = False  # 왼쪽 셀은 첫 segment에서만 헤더 출력

    while lines:
        segment_start_y = pdf.get_y()
        # 현재 페이지에 들어갈 만큼의 줄들을 출력
        while lines and pdf.get_y() + line_height <= pdf.page_break_trigger:
            pdf.set_xy(x_right, pdf.get_y())
            pdf.cell(right_width, line_height, txt=lines.pop(0), ln=1)
        segment_end_y = pdf.get_y()
        segment_height = segment_end_y - segment_start_y

        # 오른쪽 셀 영역에 테두리 그리기
        pdf.rect(x_right, segment_start_y, right_width, segment_height)

        # 왼쪽 셀: 첫 segment에는 헤더(중앙정렬) 출력, 이후는 빈 문자열로 처리
        pdf.set_xy(x_left, segment_start_y)
        if not printed_left:
            pdf.cell(left_width, segment_height, txt=left_text, align="C")  
            printed_left = True
        else:
            pdf.cell(left_width, segment_height, txt="", align="C")
        pdf.rect(x_left, segment_start_y, left_width, segment_height)

        # 만약 아직 출력할 줄이 남았다면 새 페이지 추가 후 계속 출력
        if lines:
            pdf.add_page()
    # 조사 내용 행 출력 종료 후, 다음 행을 위해 x 좌표를 재설정
    pdf.set_x(pdf.l_margin)

def draw_table_from_excel(pdf, sections, values):
    """
    엑셀의 테이블 구조를 모방하여 각 섹션(행)을 PDF에 순서대로 출력합니다.
    
    - sections 리스트 순서대로 아래 행들이 출력되어야 함:
         ["주제(키워드)", "조사 개요", "조사 내용", "출처", "알게 된 점 또는 느낀 점"]
    - "조사 내용" 행에는 오른쪽 셀에만 values["조사 내용"] (즉, "3")이 들어가고,
      나머지 행의 오른쪽 셀에는 해당하는 값이 들어갑니다.
    """
    col_widths = [50, 140]  # 왼쪽 셀, 오른쪽 셀 너비 (mm)
    line_height = 6         # 기본 줄 높이 (mm)
    
    for section in sections:
        value = values[section]
        x_left = pdf.get_x()
        y_start = pdf.get_y()
        # 왼쪽 셀: 섹션 제목 중앙정렬 출력
        pdf.multi_cell(col_widths[0], line_height, txt=section, border=0, align="C")
        y_after_left = pdf.get_y()
        left_cell_height = y_after_left - y_start

        # 오른쪽 셀: 해당 값을 출력
        pdf.set_xy(x_left + col_widths[0], y_start)
        pdf.multi_cell(col_widths[1], line_height, txt=value, border=0, align="L")
        y_after_right = pdf.get_y()
        right_cell_height = y_after_right - y_start

        row_height = max(left_cell_height, right_cell_height)
        # "알게 된 점 또는 느낀 점" 행의 경우 최소 높이를 강제
        if section == "알게 된 점 또는 느낀 점":
            base_min_height = 10 * line_height
            extra_space = 10 if pdf.page_no() >= 2 else 0
            row_height = max(row_height, base_min_height + extra_space)
        
        # 양쪽 셀에 테두리 그리기
        pdf.rect(x_left, y_start, col_widths[0], row_height)
        pdf.rect(x_left + col_widths[0], y_start, col_widths[1], row_height)
        pdf.set_xy(x_left, y_start + row_height)

def create_pdf_from_excel_template(content, template_path="drizzle_format.xlsx"):
    # 1. 엑셀 템플릿 파일 존재 여부 확인
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"엑셀 템플릿 파일이 존재하지 않습니다: {template_path}")
    
    # 2. 임시 엑셀 파일 생성
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        excel_path = tmp_excel.name
    
    # 3. 엑셀 파일 불러오기 및 템플릿에 데이터 입력
    wb = load_workbook(template_path)
    ws = wb.active
    sections = ["주제(키워드)", "조사 개요", "조사 내용", "출처", "알게 된 점 또는 느낀 점"]
    row_positions = [8, 9, 10, 20, 22]
    for idx, section in enumerate(sections):
        coord = f"B{row_positions[idx]}"
        merged_range = is_merged_cell(ws, coord)
        if merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = section
        else:
            ws[coord].value = section
    
    
    # 4. PDF 변환 시작
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_margins(5, 5, 5)
    pdf.add_page()
    
    # 5. 한글 폰트 설정 (fonts 폴더 내 H2MJRE.TTF 파일 필요)
    font_path = os.path.join("fonts", "H2MJRE.TTF")
    if not os.path.exists(font_path):
        raise Exception("fonts 폴더에서 한글 폰트를 찾을 수 없습니다. H2MJRE.TTF 또는 다른 한글 폰트가 필요합니다.")
    pdf.add_font("CustomFont", "", font_path, uni=True)
    pdf.set_font("CustomFont", size=12)
    
    pdf.cell(200, 10, txt="조사 보고서", ln=True, align="C")
    pdf.ln(5)
    

    parsed_result = parse_report_html_to_json(content)
    values = parsed_result

    print("values")
    print(values)

    
    # # 각 섹션에 넣을 값들 (테스트로 1,2,3,4,5로 대체)
    # values = {
    #     "주제(키워드)": content['주제'],
    #     "조사 개요": content['조사 개요'],
    #     "조사 내용": content['조사 내용'],
    #     "출처": content['출처'],
    #     "알게 된 점 또는 느낀 점": ""
    # }
    
    # 6. 테이블 그리기 (각 행을 순서대로 출력)
    draw_table_from_excel(pdf, sections, values)
    
    # 7. PDF 파일 저장
    pdf_path = excel_path.replace(".xlsx", ".pdf")
    pdf.output(pdf_path)
    
    return excel_path, pdf_path


def sidebar():
    """사이드바 구성"""
    load_css()
    
    # 로고 이미지
    st.sidebar.image("images/logo-removebg.png", width=128, use_container_width=True)

    # 사용자 정보 표시
    sidebar_col1, sidebar_col, sidebar_col2 = st.sidebar.columns([1.0, 0.2, 1.0]) 

    # 사용자 정보 버튼으로 변경
    sidebar_col1.markdown('<span id="button-user"></span>', unsafe_allow_html=True)
    sidebar_col1.button(
        "👋 최소정 님",
        key="user_button",
        help="사용자 프로필 메뉴",  # 툴팁 추가
        use_container_width=True
    )

    # 알림 버튼 추가
    sidebar_col2.markdown('<span id="button-alert"></span>', unsafe_allow_html=True)
    sidebar_col2.button(
        "👤 알림",
        key="alert_button",
        help="새 알림 확인",  # 툴팁 추가
        use_container_width=True  # <-- 이 부분을 True로
    )


    # 현재 페이지 상태 초기화
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "Research Report"  # 기본 페이지

    # 버튼 생성
    buttons = ["📊 탐구 질문 만들기", 
            "📝 자료 수집·탐색하기", 
            "📓 자료 분석하기", 
            "🤝 공유 노트", 
            "👨‍🎓 내 노트", 
            "🔍 탐구 라이브러리", 
            "📄 문서 서식", 
            "❔ FAQ"
            ]

    for i, button_text in enumerate(buttons):
        st.sidebar.markdown('<span id="button-manu"></span>', unsafe_allow_html=True)
        if st.sidebar.button(button_text, key=f"option_{i}"):
            st.session_state.current_page = button_text

    # 하단 버튼 그룹
    col1, col2, col3 = st.sidebar.columns([1, 0.2, 1])
    # 수정된 코드
    col1.markdown('<span id="button-manual"></span>', unsafe_allow_html=True)
    col1.button("메뉴얼")
    col3.markdown('<span id="button-logout"></span>', unsafe_allow_html=True)
    col3.button("로그아웃", key="logout", use_container_width=True)

def main_content():
    """메인 콘텐츠 영역"""
    # 레이아웃 컬럼 설정
    left_margin, main_center, right_margin = st.columns([0.01, 1.0, 0.01])
    left_margin, keyword_section, center, document_section, right_margin = st.columns(
        [0.05, 0.7, 0.05, 0.6, 0.1]
    )

    if 'editor_content' not in st.session_state:
        st.session_state.editor_content = ""

    with main_center:
        if st.session_state.current_page == "📝 자료 수집·탐색하기":
            st.subheader("자료 수집·탐색하기")
            st.write("질문 해결을 위한 자료를 수집·탐색합니다.")
            
            # 세로 구분선
            with center:
                st.markdown('<div class="vertical-line"></div>', unsafe_allow_html=True)

            # 키워드 섹션
            with keyword_section:
                keyword_col1, keyword_col2, keyword_col3, keyword_col4 = st.columns([0.2, 0.2, 0.2, 0.3])
                keyword_col1.markdown('<div class="info-cmd">추천 키워드</div>', unsafe_allow_html=True)

                keyword_col2.markdown('<span id="selectbox-grade"></span>', unsafe_allow_html=True)
                keyword_col2.selectbox(
                    "학년", 
                    ["1학년", "2학년", "3학년", "4학년", "5학년", "6학년"], 
                    label_visibility="collapsed",
                    key='selected_grade'
                )
                
                keyword_col3.markdown('<span id="selectbox-curriculum"></span>', unsafe_allow_html=True)
                keyword_col3.selectbox(
                    "교과선택", 
                    ["국어", "사회", "과학"], 
                    label_visibility="collapsed",
                    key='selected_subject'
                )
                                
                keyword_col4.text_input(
                    "키워드 입력", 
                    placeholder="키워드를 입력하세요", 
                    label_visibility="collapsed",
                    key='keyword_input'
                )

                add_vertical_space(1)
                
                # 키워드 버튼 생성
                if 'selected_text' not in st.session_state:
                    st.session_state.selected_text = None

                # RAG 결과를 담을 리스트 초기화
                if 'keywords' not in st.session_state:
                    st.session_state.keywords = []
                
                if st.session_state.keyword_input:
                    st.session_state.keywords = generate_topics(st.session_state.selected_grade, st.session_state.selected_subject, st.session_state.keyword_input)
                
                if 'keywords' in st.session_state and st.session_state.keywords:
                    for idx, keyword in enumerate(st.session_state.keywords):
                        is_selected = keyword == st.session_state.selected_text
                        if is_selected:
                            st.markdown('<span class="selected-keyword"></span>', unsafe_allow_html=True)
                        else:
                            st.markdown('<span id="keyword-button"></span>', unsafe_allow_html=True)

                        button_label = f"**{keyword}**" if is_selected else keyword
                        
                        # 콜백 함수 추가
                        def select_keyword(keyword):
                            if st.session_state.selected_text == keyword:
                                st.session_state.selected_text = None
                            else:
                                st.session_state.selected_text = keyword

                        # if st.button(button_label, key=f"kw_{idx}", help=None):
                        #     if is_selected:
                        #         st.session_state.selected_text = None
                        #     else:
                        #         st.session_state.selected_text = keyword
    #                        st.rerun()
                        # 버튼 클릭 핸들러
                        if st.button(button_label, 
                                    key=f"kw_{idx}", 
                                    help=None,
                                    # 버튼 클릭 시 부분 리렌더링 방지
                                    on_click=lambda k=keyword: select_keyword(k)):
                            pass
                    

            # 문서 편집기 섹션
            with document_section:
                if st.session_state.selected_text:
                    st.markdown(f'<div class="selected-text">{st.session_state.selected_text}</div>', unsafe_allow_html=True)
                    
                if 'editor_key' not in st.session_state:
                    st.session_state.editor_key = 0

                
                if st.session_state.editor_content is not None and isinstance(st.session_state.editor_content, dict):
                    default_text = f"""
                    <style>
                        .quill-container {{
                            margin: 0;  /* 여백 제거 */
                            padding: 0; /* 패딩 제거 */
                        }}
                    </style>
                    1. 주제
                    {st.session_state.editor_content.get('주제', '')}

                    2. 조사 개요
                    {st.session_state.editor_content.get('조사 개요', '')}

                    3. 조사 내용
                    {st.session_state.editor_content.get('조사 내용', '')}

                    4. 출처
                    {st.session_state.editor_content.get('출처', '')}
                    """
                else:
                    default_text = "편집할 내용이 없습니다. 먼저 보고서를 생성해 주세요."
                # 에디터 초기화
                content = st_quill(
                    value=default_text,
                    html=True,
                    key=f'quill_editor_{st.session_state.editor_key}'  # 동적 키 적용
                )
                
                st.session_state.editor_content = content
                
                # Create three columns to hold the buttons in one row.
                col1, col2, col3, col4, col5, col6 = st.columns([0.2, 0.5, 0.5, 0.5, 0.5, 0.2])

                with col2:
                    # Insert a marker before the button so that we can style it via CSS.
                    st.markdown('<span id="button-summary"></span>', unsafe_allow_html=True)
                    if st.button("생성", key="create_button"):
                        st.session_state.editor_content = generate_report(st.session_state.selected_grade, st.session_state.selected_subject, st.session_state.selected_text)
                        st.session_state.editor_key += 1  # 키 값 변경으로 컴포넌트 강제 리렌더링
                        st.rerun()



                with col3:
                    st.markdown('<span id="button-summary"></span>', unsafe_allow_html=True)
                    if st.button("요약", key="summary_button"):
                        if st.session_state.editor_content:
                            summary = generate_summary(st.session_state.editor_content)
                            print("_-------------------")
                            print(summary)
                            st.session_state.editor_content = summary  # 요약 결과를 반영
                            st.session_state.editor_key += 1  # 키 값 변경으로 강제 리렌더링
                            st.rerun()
                        else:
                            st.warning("요약할 내용이 없습니다. 먼저 보고서를 생성해 주세요.")


                with col4:
                    st.markdown('<span id="button-copy"></span>', unsafe_allow_html=True)
                    if st.button("복사", key="copy_button"):
                        # st_quill 위젯의 최신 내용을 직접 사용 (content 변수에 저장된 값을 활용)
                        if content:
                            # HTML 내용을 예쁜 문자열로 변환
                            formatted_text = format_copied_text(content)
#                            pyperclip.copy(formatted_text)
                            copy_to_clipboard_js(formatted_text)
                            st.success("클립보드에 복사되었습니다!")
                        else:
                            st.warning("복사할 내용이 없습니다.")

                with col5:
                    st.markdown('<span id="button-print"></span>', unsafe_allow_html=True)
                    if st.button("출력", key="print_button"):
                        
                        if st.session_state.editor_content:
                            # 엑셀 및 PDF 파일 생성
                            excel_path, pdf_path = create_pdf_from_excel_template(st.session_state.editor_content)

                            # PDF 파일 다운로드 링크 제공
                            with open(pdf_path, "rb") as f:
                                st.download_button(
                                    label="PDF 다운로드",
                                    data=f,
                                    file_name="report.pdf",
                                    mime="application/pdf"
                                )

                            # 임시 파일 삭제
                            os.remove(excel_path)
                            os.remove(pdf_path)

                            st.success("PDF 파일이 생성되었습니다. 다운로드 버튼을 클릭하세요.")
                        else:
                            st.warning("출력할 내용이 없습니다. 먼저 보고서를 생성해 주세요.")

        elif st.session_state.current_page == "📊 탐구 질문 만들기":
            st.subheader("📊 탐구 질문 만들기")
            st.write("선택한 키워드를 바탕으로 탐구 질문을 생성하고, 유형별로 분류해볼 수 있어요.")
            
            with center:
                st.markdown('<div class="vertical-line"></div>', unsafe_allow_html=True)
                
                with keyword_section:
                    keyword_col1, keyword_col2, keyword_col3, keyword_col4 = st.columns([0.2, 0.2, 0.2, 0.4])
                    keyword_col1.markdown('<div class="info-cmd">추천 키워드</div>', unsafe_allow_html=True)

                    keyword_col2.selectbox("학년", ["1학년", "2학년", "3학년", "4학년", "5학년", "6학년"], 
                                        label_visibility="collapsed", key='selected_grade')
                    keyword_col3.selectbox("교과선택", ["국어", "사회", "과학"], 
                                        label_visibility="collapsed", key='selected_subject')
                    keyword_col4.text_input("키워드 입력", placeholder="키워드를 입력하세요", 
                                            label_visibility="collapsed", key='keyword_input')
                                            
                                            add_vertical_space(1)

                if 'selected_text' not in st.session_state:
                    st.session_state.selected_text = None

                if 'keywords' not in st.session_state:
                    st.session_state.keywords = []

                if st.session_state.keyword_input:
                    st.session_state.keywords = generate_topics(
                        st.session_state.selected_grade, 
                        st.session_state.selected_subject, 
                        st.session_state.keyword_input
                    )

                if st.session_state.keywords:
                    for idx, keyword in enumerate(st.session_state.keywords):
                        is_selected = keyword == st.session_state.selected_text
                        style_span = '<span class="selected-keyword"></span>' if is_selected else '<span id="keyword-button"></span>'
                        st.markdown(style_span, unsafe_allow_html=True)
                        button_label = f"**{keyword}**" if is_selected else keyword

                        def select_keyword(k=keyword):
                            st.session_state.selected_text = None if st.session_state.selected_text == k else k

                        if st.button(button_label, key=f"kwq_{idx}", on_click=select_keyword):
                            pass

                with document_section:
                    if st.session_state.selected_text:
                        st.markdown(f'<div class="selected-text">{st.session_state.selected_text}</div>', unsafe_allow_html=True)

                    if 'generated_questions' not in st.session_state:
                        st.session_state.generated_questions = {}

                    col1, col2 = st.columns([0.5, 0.5])

                    with col1:
                        if st.button("질문 생성"):
                            # 질문 생성
                            result = generate_question(
                                st.session_state.selected_grade,
                                st.session_state.selected_subject,
                                st.session_state.selected_text
                            )
                            st.session_state.generated_questions = result
                            st.success("질문이 생성되었습니다!")

                    with col2:
                        if st.session_state.generated_questions:
                            copy_text = ""
                            for key, questions in st.session_state.generated_questions.items():
                                copy_text += f"【{key}】\n" + "\n".join(f"- {q}" for q in questions) + "\n\n"

                            if st.button("복사"):
                                copy_to_clipboard_js(copy_text)
                                st.success("질문이 복사되었습니다!")

                    st.divider()

                    # 표로 출력
                    if st.session_state.generated_questions:
                        st.write("### 생성된 탐구 질문")

                        max_len = max(len(qs) for qs in st.session_state.generated_questions.values())
                        data = {
                            "사실적 질문": st.session_state.generated_questions.get("사실적 질문", []),
                            "개념적 질문": st.session_state.generated_questions.get("개념적 질문", []),
                            "논쟁적 질문": st.session_state.generated_questions.get("논쟁적 질문", [])
                        }

                        rows = []
                        for i in range(max_len):
                            row = {
                                "사실적 질문": data["사실적 질문"][i] if i < len(data["사실적 질문"]) else "",
                                "개념적 질문": data["개념적 질문"][i] if i < len(data["개념적 질문"]) else "",
                                "논쟁적 질문": data["논쟁적 질문"][i] if i < len(data["논쟁적 질문"]) else "",
                            }
                            rows.append(row)

                        df = pd.DataFrame(rows)
                        st.dataframe(df, use_container_width=True, hide_index=True)
                    else:
                        st.info("먼저 키워드를 선택하고 '질문 생성' 버튼을 눌러주세요.")
                        
        elif st.session_state.current_page == "📓 내 노트":
            st.subheader("내 노트")
            st.write("개인 노트 관리 기능이 준비 중입니다.")

       
        elif st.session_state.current_page == "📓 내 노트":
            st.subheader("내 노트")
            st.write("개인 노트 관리 기능이 준비 중입니다.")

def main():
    """메인 앱 설정"""
    st.set_page_config(
        page_title="AI 활용 질문 기반 수업",
        page_icon="🔎",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    
    sidebar()
    main_content()

if __name__ == '__main__':
    main()
